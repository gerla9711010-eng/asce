# gui_main.py - 不動產售屋表自動填寫工具
# 執行方式：python gui_main.py 或雙擊 啟動工具.vbs

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import os
import re
import sys
import datetime

# ── 路徑設定 ──────────────────────────────────────────
BASE_DIR      = os.path.dirname(os.path.abspath(__file__))
TEMPLATE      = os.path.join(BASE_DIR, 'template', 'sale_template.xltx')
TEMPLATE_LAND = os.path.join(BASE_DIR, 'template', 'land_template.xltx')
OUTPUT_DIR    = os.path.join(BASE_DIR, 'output')
os.makedirs(OUTPUT_DIR, exist_ok=True)

sys.path.insert(0, BASE_DIR)
from parser import parse_land, parse_building, merge, _to_ping, _share_to_float, is_full_ownership
from bot_104 import Bot104, fetch_zoning
from confirm_wizard import ConfirmWizard, _build_land_steps


def _clean_floor_label(label: str) -> str:
    """謄本『層次』欄位有些格式是『突出物一層：２５．７２平方公尺』，
    後面那段是面積重複帶出來的，不是樓層名稱，填表只要冒號前面那段。"""
    return re.split(r'[：:]', label, maxsplit=1)[0].strip()


# ─────────────────────────────────────────────────────
#  填表邏輯
#  - 直接讀 .xltx 範本（openpyxl），不需 LibreOffice
#  - 只寫指定資料格；選項方格填黃色；函數格(F3 總建坪 / W3 單價)完全不碰
#  - 儲存格座標依範本實際合併格對位（2026-05 重校）
# ─────────────────────────────────────────────────────
def fill_excel(data: dict, output_path: str, is_rental: bool = False, log=print):
    """填售屋表 / 租賃表。
    租賃表與售屋表同一張範本，唯一差別：把「總價款:」標籤換成「月租:」，
    金額仍填 N3（單價 W3=TRUNC(N3/F3) 自動變成月租/坪）。其餘欄位完全一樣。
    log：填表過程有需要人工核對的狀況（欄位放不下）時用來提醒，預設 print；
    GUI 呼叫時應傳 self._log，不然警告會印到看不到的地方，等於沒提醒。"""
    from openpyxl import load_workbook
    from openpyxl.utils import coordinate_to_tuple
    from openpyxl.styles import PatternFill

    YELLOW = PatternFill(fill_type='solid', fgColor='FFFF00')

    wb = load_workbook(TEMPLATE)
    wb.template = False   # 範本是 .xltx，不關掉這旗標存出的 .xlsx 內部類型會是 template.main+xml，別台電腦 Excel 拒開
    ws = wb.active

    if is_rental:
        ws['L3'] = '月 租:'   # 售屋表的「總價款:」→ 租賃改成「月租:」（金額同樣填 N3）

    def put(addr, value):
        """寫值到合併格左上角"""
        if value is None or value == '':
            return
        ws[addr] = value

    def fill(addr):
        """把選項方格（含整個合併範圍）填滿黃色"""
        if not addr:
            return
        r, c = coordinate_to_tuple(addr)
        for mc in ws.merged_cells.ranges:
            if mc.min_row <= r <= mc.max_row and mc.min_col <= c <= mc.max_col:
                for rr in range(mc.min_row, mc.max_row + 1):
                    for cc in range(mc.min_col, mc.max_col + 1):
                        ws.cell(rr, cc).fill = YELLOW
                return
        ws[addr].fill = YELLOW

    # ── 標頭 ──
    put('A1',  data.get('case_name'))      # 案名
    put('P1',  data.get('address'))        # 物件座落
    put('AS1', data.get('builder'))        # 建設公司
    put('AS2', data.get('building_name'))  # 大樓名稱
    put('N3',  data.get('price'))          # 總價款（W3 單價自動算，勿填）

    # ── 設定（抵押）──
    if data.get('mortgage'):
        fill('AJ4'); put('AL3', data.get('mortgage_amount')); put('AR3', data.get('mortgage_bank'))
    else:
        fill('AG4')

    # ── 坪數（F3 總建坪 = SUM(F7:J31) 自動算，勿填）──
    # 主建物「室內」下面有 5 格（E12/F12、E14/F14…E20/F20；E欄範本預設是
    # 「室內」／「F」）：謄本記載幾層就分別填幾格，不要全部擠進 F12 一格
    # （超過 5 筆這個範本放不下，全數併進最後一格，並在 log 提醒要手動拆）
    floor_pings = data.get('floor_pings') or []
    floor_slots = ['F12', 'F14', 'F16', 'F18', 'F20']
    label_slots = ['E12', 'E14', 'E16', 'E18', 'E20']
    if floor_pings:
        overflow = floor_pings[len(floor_slots) - 1:]
        shown = floor_pings[:len(floor_slots) - 1] if len(floor_pings) > len(floor_slots) else floor_pings
        for addr, (_, ping) in zip(floor_slots, shown):
            put(addr, ping)
        # 只有真的有多層時才把 E 欄改寫成謄本實際層次名稱（一層/二層/突出物一層…）；
        # 一般大樓單一樓層戶只有1筆，E12 維持範本預設「室內」，不要被謄本樓層字樣蓋掉。
        if len(floor_pings) > 1:
            for addr, (label, _) in zip(label_slots, shown):
                put(addr, _clean_floor_label(label))
        if len(floor_pings) > len(floor_slots):
            put(floor_slots[len(shown)], round(sum(p for _, p in overflow), 2))
            put(label_slots[len(shown)], '其他樓層合計')
            log(f'⚠ 主建物層數({len(floor_pings)})超過範本5格，最後一格已合併{len(overflow)}筆，請人工核對拆分')
    else:
        put('F12', data.get('area_indoor'))    # 室內（沒有逐層資料時退回總數）

    put('F22', data.get('area_balcony'))   # 陽台
    put('F24', data.get('area_canopy'))    # 雨遮/花台

    # 附屬建物裡謄本上有、但不是陽台/雨遮/花台的類型 → 填進「其他-」欄（只有一格，
    # 超過一筆合併填同一格並在 log 提醒，避免默默漏掉某個坪數）
    extra = data.get('extra_attachments') or []
    if extra:
        labels = '、'.join(lbl for lbl, _ in extra)
        put('C26', f'其他-{labels}')
        put('F26', round(sum(p for _, p in extra), 2))
        if len(extra) > 1:
            log(f'⚠ 附屬建物有{len(extra)}種未分類類型（{labels}），已合併填同一格，請人工核對拆分')

    put('F28', data.get('area_parking'))   # 車位
    put('F30', data.get('area_common'))    # 其他公設
    put('F32', data.get('area_land'))      # 地坪
    put('AN7', data.get('road_width'))     # 面前道路

    # ── 建築型態 + 樓層（（最低/最高）樓，中間格填 104 的 floor_low_high）──
    bt = data.get('building_type')
    fill({'透天': 'N8', '大樓': 'N10', '公寓': 'N12', '其他': 'N14'}.get(bt))
    flh = data.get('floor_low_high')   # 例如 B4/24，來自 104
    if flh:
        put({'透天': 'Q8', '大樓': 'Q10', '公寓': 'Q12'}.get(bt, 'Q10'), flh)

    # ── 隔局 ──
    put('X8',  data.get('layout_rooms'))
    put('AB8', data.get('layout_halls'))
    put('AG8', data.get('layout_baths'))

    # ── 主要結構 ──
    s = data.get('structure', '')
    for k, v in {'加強磚造': 'N16', '鋼骨': 'N18', '鋼筋混凝土': 'N20', '其它': 'R16'}.items():
        if k in s:
            fill(v); break

    # ── 警衛 ──
    fill({'無': 'Y14', '日': 'AA14', '24H': 'AD14'}.get(data.get('guard')))

    # ── 座向 ──
    put('AA12', data.get('window_facing'))  # 落地窗朝
    put('AI12', data.get('door_facing'))    # 大門朝

    # ── 車位 ──
    if data.get('_parking_yn') == '有':
        fill('AN14')                       # 車位:有
        fill('AU14')                       # 公設內含（謄本車位登記於共有部分）
        # 地上 / 地下
        if data.get('_parking_pos') == '地上':
            fill('AP18')
        elif data.get('_parking_pos') == '地下':
            fill('AR18')
        put('AS18', data.get('_parking_floor'))   # 第幾層
        put('AV18', data.get('parking_no'))       # 編號
        # 平面 / 機械
        pt = data.get('_parking_type')
        if pt == '平面':
            fill('AN20')
        elif pt == '機械':
            fill('AT20')
            mech = data.get('_parking_mech')
            fill({'上': 'AN22', '中': 'AP22',
                  '下': 'AR22', '橫移': 'AT22'}.get(mech))

    # ── 入口型式 ──
    ent = data.get('_entrance_type')
    if ent == '坡道':
        fill('AN24')
    elif ent == '升降':
        fill('AT24')

    # ── 管理費 ──
    put('X16', data.get('mgmt_fee'))

    # ── 電梯 / 學區 ──
    put('X18',  data.get('elevator_units'))
    put('AD18', data.get('elevator_count'))
    put('X20',  data.get('school_junior'))
    put('AD20', data.get('school_primary'))

    # ── 完工日 ──
    put('O24', data.get('complete_year'))
    put('Q24', data.get('complete_month'))
    put('S24', data.get('complete_day'))

    # ── 周邊 / 總戶數 ──
    put('X22', data.get('market_nearby'))
    put('AG22', data.get('park_nearby'))
    put('X24', data.get('mrt_nearby'))
    put('AG24', data.get('moto_parking'))
    put('N26', data.get('total_units'))
    fill({'空屋': 'Y26', '自住': 'AC26', '租賃': 'AG26'}.get(data.get('current_status')))

    # ── 所有權 ──
    if data.get('ownership') == '全部':
        fill('B36')
    else:
        fill('B38'); put('H38', data.get('land_share'))

    # ── 使用用途 ──
    up = data.get('usage_purpose', '')
    if any(k in up for k in ['住宅', '住家', 'H1', 'H2', '集合住宅']):
        fill('P22')
    elif any(k in up for k in ['店舖', '商業', '店面']):
        fill('N22')
    elif any(k in up for k in ['辦公', '事務所']):
        fill('S22')

    # ── 使用區分（由網站查詢補入）──
    #   住宅/商業/工業 → 勾大類塗黃 + 種別填專屬格，不寫 G44
    #   其他（市場用地等）→ 整串文字寫 G44 並塗黃
    uz = data.get('usage_zone') or ''
    sz = data.get('special_zone', '') or ''
    if uz not in ('住宅區', '商業區', '工業區'):   # 沒大類就從文字推斷
        if '商業區' in sz:   uz = '商業區'
        elif '工業區' in sz: uz = '工業區'
        elif '住宅區' in sz: uz = '住宅區'

    if uz in ('住宅區', '商業區', '工業區'):
        fill({'住宅區': 'B42', '商業區': 'B44', '工業區': 'G42'}[uz])
        if uz == '住宅區':
            m = re.search(r'([一二三四五六七八九十]+)\s*種', sz)
            if m: put('E42', m.group(1))          # 第「三」種住宅區
        elif uz == '商業區':
            m = re.search(r'([一二三四五六七八九十]+)\s*種', sz)
            if m: put('E44', m.group(1))          # 第「四」種商業區
        else:                                     # 工業區
            kind = re.sub(r'業區$', '', sz).strip()
            if kind: put('K42', kind)             # 乙種工業區 → 「乙種工」
    elif sz:
        put('G44', sz)
        fill('G44')

    # ── 公告現值 ──
    put('E46', data.get('land_announcement'))

    # ── 訴求重點（最多 5 點）──
    for cell, pt in zip(['AL30', 'AL32', 'AL34', 'AL36', 'AL38'],
                        data.get('selling_points', [])):
        put(cell, pt)

    # ── 專員 ──
    put('AO40', data.get('agent_name', '薛力瑜'))
    put('AO42', data.get('agent_name2', '周珈伊'))

    wb.save(output_path)
    return output_path


# ─────────────────────────────────────────────────────
#  土地表填寫（獨立版型，與售屋/租賃不同範本）
#  規則同售屋表：值填標籤右格、打勾格填黃色、函數格不碰
#  座標依 land_template.xltx 實際合併格對位
# ─────────────────────────────────────────────────────
def fill_land(data: dict, output_path: str, is_rental: bool = False):
    from openpyxl import load_workbook
    from openpyxl.utils import coordinate_to_tuple
    from openpyxl.styles import PatternFill

    YELLOW = PatternFill(fill_type='solid', fgColor='FFFF00')

    wb = load_workbook(TEMPLATE_LAND)
    wb.template = False   # 同售屋表：關掉範本旗標，別台電腦才開得了
    ws = wb.active

    def put(addr, value):
        if value is None or value == '':
            return
        ws[addr] = value

    def fill(addr):
        if not addr:
            return
        r, c = coordinate_to_tuple(addr)
        for mc in ws.merged_cells.ranges:
            if mc.min_row <= r <= mc.max_row and mc.min_col <= c <= mc.max_col:
                for rr in range(mc.min_row, mc.max_row + 1):
                    for cc in range(mc.min_col, mc.max_col + 1):
                        ws.cell(rr, cc).fill = YELLOW
                return
        ws[addr].fill = YELLOW

    # ── 標頭 ──
    put('D2', data.get('case_name'))       # 案名
    put('L2', data.get('address'))         # 物件座落（土地：地段地號）
    put('AL2', data.get('key_no'))         # 鑰匙編號

    # ── 地坪 ──
    put('E4', data.get('area_land'))       # 地坪（坪）

    # ── 總價款 / 租金 ──
    if is_rental:
        put('J9', data.get('price'))       # 租金（萬）
        put('J14', data.get('deposit'))    # 押金（萬）
    else:
        put('J4', data.get('price'))       # 總價款（萬）

    # ── 貸款（他項權利）──
    if data.get('mortgage'):
        fill('Z5')                         # 有
        put('AB4', data.get('mortgage_amount'))
        put('AG4', data.get('mortgage_bank'))
    else:
        fill('V5')                         # 無

    # ── 面前道路 ──
    put('AG8', data.get('road_width'))

    # ── 建蔽率 / 容積率 / 寬度 / 長度（外部來源或手動）──
    put('E9',  data.get('coverage_ratio'))   # 建蔽率 %
    put('E13', data.get('floor_ratio'))      # 容積率 %
    put('E21', data.get('land_width'))       # 寬度 米
    put('E25', data.get('land_depth'))       # 長度 米

    # ── 土地分區（使用分區查詢補入）──
    uz = data.get('usage_zone') or ''
    sz = data.get('special_zone') or ''
    put('E17', sz or uz)                     # 完整分區文字，例：第三種住宅區

    # ── 用途（謄本使用地類別，預設建地）──
    put('J21', data.get('usage_type') or '建地')

    # ── 地上建物：有才填隔局（房/廳/衛）──
    if data.get('_has_building') == '有':
        put('S9',  data.get('lot_rooms'))
        put('W9',  data.get('lot_halls'))
        put('AA9', data.get('lot_baths'))

    # ── 現況（空地 / 建物 / 租賃）──
    cur = data.get('current_status') or ('租賃' if is_rental else '空地')
    fill({'空地': 'T25', '建物': 'W25', '租賃': 'AA25'}.get(cur))

    # ── 所有權 全部 / 持分（此表用實心■/空心□，非黃底）──
    if data.get('ownership') == '全部':
        ws['A31'] = '■'; ws['E31'] = '□'
    else:
        ws['A31'] = '□'; ws['E31'] = '■'

    # ── 訴求重點（最多 9 欄，確認視窗逐欄輸入）──
    for cell, pt in zip(['AD29', 'AD31', 'AD33', 'AD35', 'AD37',
                         'AD39', 'AD41', 'AD43', 'AD45'],
                        data.get('selling_points', [])):
        put(cell, pt)

    # ── 專員 ──
    put('D39', data.get('agent_name', '薛力瑜'))
    put('D41', data.get('agent_phone'))

    wb.save(output_path)
    return output_path


# ─────────────────────────────────────────────────────
#  數字工具（坪數加總用）
# ─────────────────────────────────────────────────────
def _to_num(v):
    """把坪數值轉成 float；None / 抓不到數字 → None。支援 '12.34'、'12.34坪'。"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    m = re.search(r'-?\d+(?:\.\d+)?', str(v))
    return float(m.group()) if m else None


# ─────────────────────────────────────────────────────
#  GUI
# ─────────────────────────────────────────────────────
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title('不動產售屋表自動填寫工具')
        self.resizable(True, True)
        self.geometry('600x600')
        self.bot = None           # Bot104 instance
        self.data_104 = None      # 104 抓到的社區資料
        self._build_ui()

    # ── UI 建立 ──────────────────────────────
    def _build_ui(self):
        pad = dict(padx=12, pady=6)

        # 表格類型（一開始手動選：決定走哪條填表分支）
        frame_type = ttk.LabelFrame(self, text='① 先選表格類型', padding=8)
        frame_type.pack(fill='x', **pad)

        self.trans_var     = tk.StringVar(value='買賣')   # 買賣 / 租賃
        self.obj_kind_var  = tk.StringVar(value='建物')   # 建物 / 土地
        self.bldg_type_var = tk.StringVar(value='大樓')   # 大樓 / 透天 / 公寓
        self._DEF_BG = self.cget('bg')                    # 未選中時的底色

        # 第一行：買賣 / 租賃
        ttk.Label(frame_type, text='交易類型：').grid(row=0, column=0, sticky='w', pady=2)
        row_t = ttk.Frame(frame_type); row_t.grid(row=0, column=1, sticky='w')
        self._trans_btns = {}
        for v in ('買賣', '租賃'):
            b = tk.Button(row_t, text=v, width=8, relief='ridge', bd=2,
                          command=lambda x=v: self._select_trans(x))
            b.pack(side='left', padx=3)
            self._trans_btns[v] = b

        # 第二行：建物 / 土地（＋建物時的型態下拉）
        ttk.Label(frame_type, text='物件類型：').grid(row=1, column=0, sticky='w', pady=2)
        row_o = ttk.Frame(frame_type); row_o.grid(row=1, column=1, sticky='w')
        self._obj_btns = {}
        for v in ('建物', '土地'):
            b = tk.Button(row_o, text=v, width=8, relief='ridge', bd=2,
                          command=lambda x=v: self._select_obj(x))
            b.pack(side='left', padx=3)
            self._obj_btns[v] = b
        self.cb_bldg_type = ttk.Combobox(row_o, textvariable=self.bldg_type_var,
                                         values=['大樓', '透天', '公寓'],
                                         state='readonly', width=8)
        self.cb_bldg_type.pack(side='left', padx=(12, 0))

        self._select_trans('買賣')   # 套用預設高亮
        self._select_obj('建物')

        # 謄本選擇（多持分：土地/建物都可按 ＋新增 加列，坪數會自動加總）
        frame_pdf = ttk.LabelFrame(self, text='謄本 PDF（多持分可按 ＋新增，坪數自動加總）', padding=8)
        frame_pdf.pack(fill='x', **pad)

        self.land_vars = []       # list[StringVar] 土地謄本
        self.bldg_vars = []       # list[StringVar] 建物謄本

        ttk.Label(frame_pdf, text='土地謄本').grid(row=0, column=0, sticky='w')
        self.land_rows = ttk.Frame(frame_pdf)
        self.land_rows.grid(row=1, column=0, sticky='w')
        ttk.Button(frame_pdf, text='＋ 新增土地謄本',
                   command=lambda: self._add_pdf_row('land')).grid(row=2, column=0, sticky='w', pady=(0, 8))

        ttk.Label(frame_pdf, text='建物謄本').grid(row=3, column=0, sticky='w')
        self.bldg_rows = ttk.Frame(frame_pdf)
        self.bldg_rows.grid(row=4, column=0, sticky='w')
        ttk.Button(frame_pdf, text='＋ 新增建物謄本',
                   command=lambda: self._add_pdf_row('bldg')).grid(row=5, column=0, sticky='w')

        self._add_pdf_row('land')     # 預設各一列
        self._add_pdf_row('bldg')

        # 104 自動化區塊
        frame_104 = ttk.LabelFrame(self, text='104 自動查詢（選用）', padding=8)
        frame_104.pack(fill='x', **pad)

        row1 = ttk.Frame(frame_104); row1.pack(anchor='w', fill='x')
        self.btn_104_open = ttk.Button(row1, text='🌐 開啟 104（自動登入）',
                                       command=self._open_104)
        self.btn_104_open.pack(side='left')
        self.btn_104_go = ttk.Button(row1, text='✅ 完成登入，自動產出',
                                     command=self._login_done_and_run, state='disabled')
        self.btn_104_go.pack(side='left', padx=8)

        self.lbl_104_status = ttk.Label(frame_104, text='（未啟動 104）', foreground='#888')
        self.lbl_104_status.pack(anchor='w', pady=(4, 0))

        # 輸出路徑
        frame_out = ttk.LabelFrame(self, text='輸出資料夾', padding=8)
        frame_out.pack(fill='x', **pad)
        self.out_var = tk.StringVar(value=OUTPUT_DIR)
        ttk.Entry(frame_out, textvariable=self.out_var, width=50).pack(side='left', padx=(0, 4))
        ttk.Button(frame_out, text='選擇', width=6,
                   command=self._pick_out).pack(side='left')

        # 執行按鈕
        self.btn_run = ttk.Button(self, text='▶  開始產出售屋表', command=self._run)
        self.btn_run.pack(pady=8)

        # 進度 / 日誌
        frame_log = ttk.LabelFrame(self, text='執行記錄', padding=4)
        frame_log.pack(fill='both', expand=True, padx=12, pady=(0, 12))
        self.log = tk.Text(frame_log, height=8, state='disabled',
                           font=('Consolas', 9), bg='#1e1e1e', fg='#d4d4d4')
        self.log.pack(fill='both', expand=True)

    # ── 謄本列（可多筆持分）────────────────────
    def _add_pdf_row(self, kind):
        vars_list = self.land_vars if kind == 'land' else self.bldg_vars
        parent    = self.land_rows if kind == 'land' else self.bldg_rows
        var = tk.StringVar()
        vars_list.append(var)
        row = ttk.Frame(parent)
        row.pack(anchor='w', pady=1)
        ttk.Entry(row, textvariable=var, width=42).pack(side='left', padx=(0, 2))
        ttk.Button(row, text='選擇', width=5,
                   command=lambda v=var: self._pick_pdf(v)).pack(side='left')
        ttk.Button(row, text='✕', width=2,
                   command=lambda: self._remove_pdf_row(kind, row, var)).pack(side='left', padx=(2, 0))

    def _remove_pdf_row(self, kind, row, var):
        vars_list = self.land_vars if kind == 'land' else self.bldg_vars
        if len(vars_list) <= 1:      # 至少保留一列，只清空
            var.set('')
            return
        if var in vars_list:
            vars_list.remove(var)
        row.destroy()
        self.data_104 = None

    def _land_paths(self):
        return [v.get().strip() for v in self.land_vars if v.get().strip()]

    def _bldg_paths(self):
        return [v.get().strip() for v in self.bldg_vars if v.get().strip()]

    # ── 表格類型選擇（螢光黃高亮）────────────
    def _select_trans(self, value):
        self.trans_var.set(value)
        self._highlight(self._trans_btns, value)

    def _select_obj(self, value):
        self.obj_kind_var.set(value)
        self._highlight(self._obj_btns, value)
        if value == '建物':          # 建物才顯示型態下拉（大樓/透天/公寓）
            self.cb_bldg_type.pack(side='left', padx=(12, 0))
        else:                        # 土地：收起下拉，走土地流程
            self.cb_bldg_type.pack_forget()

    def _highlight(self, btns, selected):
        YELLOW = '#FFF200'           # 螢光黃
        for v, b in btns.items():
            if v == selected:
                b.config(bg=YELLOW, activebackground=YELLOW, relief='solid',
                         font=('TkDefaultFont', 9, 'bold'))
            else:
                b.config(bg=self._DEF_BG, activebackground=self._DEF_BG,
                         relief='ridge', font=('TkDefaultFont', 9, 'normal'))

    # ── 表格類型判斷 ────────────────────────
    def _is_rental(self):
        return self.trans_var.get() == '租賃'

    def _is_land(self):
        return self.obj_kind_var.get() == '土地'

    def _land_not_ready(self):
        """104 查詢只適用建物案；土地案不需要 104，擋下並提示。"""
        if self._is_land():
            messagebox.showinfo(
                '土地案免 104',
                '土地案不需要 104 社區查詢。\n'
                '請直接選好土地謄本，按「開始產出售屋表」即可。')
            return True
        return False

    # ── 檔案選擇 ────────────────────────────
    def _pick_pdf(self, var):
        p = filedialog.askopenfilename(filetypes=[('PDF 檔案', '*.pdf')])
        if p:
            var.set(p)
            self.data_104 = None      # 換謄本 → 清掉舊的 104 資料

    def _pick_out(self):
        d = filedialog.askdirectory(initialdir=self.out_var.get())
        if d:
            self.out_var.set(d)

    # ── 開啟 104 給使用者登入 ────────────────
    def _open_104(self):
        if self._land_not_ready():
            return
        if not self._land_paths() or not self._bldg_paths():
            messagebox.showwarning('請先選謄本',
                '請先選土地與建物謄本 PDF，再開 104。\n'
                '（程式會用第 1 筆建物謄本的門牌去搜尋 104）')
            return
        self.btn_104_open.config(state='disabled')
        self.lbl_104_status.config(text='⏳ 啟動瀏覽器…', foreground='#888')

        def worker():
            try:
                self.bot = Bot104(log=self._log)
                self.bot.open_login()
                self.lbl_104_status.config(
                    text='✅ 已自動登入 104，請按右邊「完成登入，自動產出」',
                    foreground='#2a7')
                self.btn_104_go.config(state='normal')
            except Exception as e:
                self._log(f'❌ 開啟 104 失敗：{e}')
                self.lbl_104_status.config(text=f'❌ 啟動失敗：{e}', foreground='#c33')
                self.btn_104_open.config(state='normal')
        threading.Thread(target=worker, daemon=True).start()

    # ── 登入完成 → 自動搜尋+抓資料+產出 ──
    def _login_done_and_run(self):
        if not self.bot:
            return
        self.btn_104_go.config(state='disabled')
        self.btn_run.config(state='disabled')
        threading.Thread(target=self._worker_with_104, daemon=True).start()

    def _worker_with_104(self):
        try:
            self._log('── 解析謄本 ...')
            data, land = self._combine_parcels(self._land_paths(), self._bldg_paths())
            data['building_type'] = self.bldg_type_var.get()   # 手動選的建物型態，蓋掉自動判斷
            self._log(f'  門牌：{data.get("address")}（建物型態：{data["building_type"]}）')

            # 用門牌跑 104
            self._log('── 自動搜尋 104 …')
            d104 = self.bot.search_and_fetch(data.get('address', ''))
            if d104:
                self._apply_104(data, d104)
                self.data_104 = d104
            else:
                self._log('  → 跳過 104，僅以謄本資料輸出')

            # 查使用分區（用 104 的瀏覽器開新分頁）
            self._log('── 查詢使用分區 …')
            self._apply_zoning(data, land, driver=self.bot.driver)

            # 主執行緒跳 wizard 確認 → 填表
            self.after(0, lambda d=data: self._wizard_and_produce(d, close_bot=True))
        except Exception as e:
            import traceback
            self._log(f'❌ 錯誤：{e}')
            self._log(traceback.format_exc())
            self.after(0, lambda: self._reset_after_run(close_bot=True))

    # ── wizard 確認 + 產出（main thread）──
    def _wizard_and_produce(self, data: dict, close_bot: bool = False):
        try:
            self._log('── 跳出確認視窗，逐項確認 ...')
            wiz = ConfirmWizard(self, data, log=self._log)
            result = wiz.run()
            if result is None:
                return
            self._produce(result)
        except Exception as e:
            import traceback
            self._log(f'❌ 錯誤：{e}')
            self._log(traceback.format_exc())
        finally:
            self._reset_after_run(close_bot=close_bot)

    def _reset_after_run(self, close_bot: bool = False):
        self.btn_run.config(state='normal')
        self.btn_104_open.config(state='normal')
        self.data_104 = None          # 清掉本次 104 資料，避免殘留套到下一筆
        self.btn_104_go.config(state='disabled')
        if close_bot:
            try:
                if self.bot:
                    self.bot.close()
            except Exception:
                pass
            self.bot = None
            self.lbl_104_status.config(text='（已關閉 104 視窗）', foreground='#888')

    def _produce(self, data: dict):
        is_rental = self._is_rental()
        kind = '租賃表' if is_rental else '售屋表'
        self._log(f'── 填寫{kind} ...')
        ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        addr_short = (data.get('address') or '物件').replace('高雄市', '').replace('屏東縣', '')[:20]
        out_name = f"{kind}_{addr_short}_{ts}.xlsx"
        out_path = os.path.join(self.out_var.get(), out_name)
        fill_excel(data, out_path, is_rental=is_rental, log=self._log)
        self._log(f'✅ 完成！輸出：{out_path}')
        self._log('──────────────────────────────')
        if sys.platform == 'win32':
            os.startfile(os.path.dirname(out_path))

    @staticmethod
    def _apply_104(data: dict, d104: dict):
        """把 104 欄位套進填表 data（只覆蓋有值的）"""
        m = {
            'building_name': d104.get('community_name'),
            'builder':       d104.get('builder'),
            'total_units':   d104.get('total_units'),
            'school_primary':d104.get('school_primary'),
            'school_junior': d104.get('school_junior'),
            'special_zone':  d104.get('special_zone'),
            'guard':         d104.get('guard'),
            'floor_low_high':d104.get('floor_low_high'),
        }
        for k, v in m.items():
            if v not in (None, '', []):
                data[k] = v

    # 建物坪數欄位（跨多建號時一起加總）
    _BLDG_AREA_FIELDS = ('area_indoor', 'area_balcony', 'area_canopy',
                         'area_parking', 'area_common')

    def _combine_parcels(self, land_paths, bldg_paths):
        """解析並合併多筆土地/建物謄本。
        地坪   = 所有土地謄本地坪加總（多持分才會是完整地坪）
        建物坪 = 所有建物謄本坪數加總（多建號：主建物＋增建＋車位建號…）
        其餘欄位（門牌/格局/樓層/社區…）取第 1 筆土地 + 第 1 筆建物。
        回傳 (data, land0)；land0 供使用分區查詢用（以第 1 筆地號為準）。
        """
        land0 = parse_land(land_paths[0])
        bldg0 = parse_building(bldg_paths[0])
        data = merge(land0, bldg0)

        # ── 地坪加總 + 持分加總是否等於全部 ──
        # 每張土地謄本的 area_land 已經是「這個所有權人的持分坪數」（merge() 用
        # land_area × land_share 算的），所以加總才會是完整地坪；但舊版只加坪數，
        # 沒有同步檢查持分加起來是不是剛好=1——結果 3 個共有人謄本都餵了、地坪也
        # 加對了，所有權卻還是照 land0（第1筆）標「持分 3分之1」，沒有改標「全部」。
        if len(land_paths) > 1:
            total, share_total, got = 0.0, 0.0, False
            for i, lp in enumerate(land_paths):
                di = data if i == 0 else merge(parse_land(lp), bldg0)
                v = _to_num(di.get('area_land'))
                self._log(f'   土地{i + 1} 地坪：{di.get("area_land")}（持分 {di.get("land_share")}）')
                if v is not None:
                    total += v; got = True
                if di.get('land_share'):
                    share_total += _share_to_float(di['land_share'])
            if got:
                data['area_land'] = round(total, 2)
                self._log(f'  ✓ 地坪加總 {len(land_paths)} 筆 = {data["area_land"]} 坪')
            if abs(share_total - 1.0) < 0.01:
                data['ownership'] = '全部'
                self._log(f'  ✓ 持分加總 ≈ {share_total:.3f}（等於全部），所有權改標「全部」而非「持分」')

        # ── 建物坪數加總（同一間只算一次）──
        #   建物標示部面積是「整間實際面積」，不隨持分切分：
        #   多所有權人各放一張建物謄本時，面積相同 → 判定同一建號，只算一次；
        #   增建 / 車位等不同建號面積不同 → 照加。
        if len(bldg_paths) > 1:
            sums = {k: 0.0 for k in self._BLDG_AREA_FIELDS}
            got = {k: False for k in self._BLDG_AREA_FIELDS}
            seen, counted, dup = set(), 0, 0
            for j, bp in enumerate(bldg_paths):
                dj = data if j == 0 else merge(land0, parse_building(bp))
                sig = tuple(round(_to_num(dj.get(k)) or 0.0, 2)
                            for k in self._BLDG_AREA_FIELDS)
                if sig in seen:
                    dup += 1
                    self._log(f'   建物{j + 1}：面積與前面相同 → 同一建號（多持分），不重複加')
                    continue
                seen.add(sig); counted += 1
                for k in self._BLDG_AREA_FIELDS:
                    v = _to_num(dj.get(k))
                    if v is not None:
                        sums[k] += v; got[k] = True
            for k in self._BLDG_AREA_FIELDS:
                if got[k]:
                    data[k] = round(sums[k], 2)
            self._log(f'  ✓ 建物坪數：{counted} 個不同建號加總'
                      + (f'（{dup} 張為同一間持分，已略過）' if dup else ''))

        return data, land0

    def _apply_zoning(self, data: dict, land: dict, driver=None):
        """查高雄市使用分區，覆蓋 usage_zone / special_zone（官網最準）。"""
        try:
            z = fetch_zoning(land.get('district'), land.get('section'),
                             land.get('land_no'), driver=driver, log=self._log)
        except Exception as e:
            self._log(f'⚠ 使用分區查詢例外：{e}')
            return
        if not z:
            return
        if z.get('usage_zone'):
            data['usage_zone'] = z['usage_zone']
        if z.get('special_zone'):
            data['special_zone'] = z['special_zone']

    # ── 日誌輸出 ────────────────────────────
    def _log(self, msg: str):
        if threading.current_thread() is threading.main_thread():
            self._log_ui(msg)
        else:
            self.after(0, self._log_ui, msg)

    def _log_ui(self, msg: str):
        self.log.config(state='normal')
        self.log.insert('end', msg + '\n')
        self.log.see('end')
        self.log.config(state='disabled')

    # ── 主流程 ──────────────────────────────
    def _run(self):
        land_paths = self._land_paths()

        # ── 土地案：只需土地謄本，走土地分支 ──
        if self._is_land():
            if not land_paths:
                messagebox.showerror('錯誤', '請先選擇土地謄本 PDF')
                return
            self.btn_run.config(state='disabled')
            threading.Thread(target=self._worker_land,
                             args=(land_paths,), daemon=True).start()
            return

        # ── 建物案：土地 + 建物謄本 ──
        bldg_paths = self._bldg_paths()
        if not land_paths or not bldg_paths:
            messagebox.showerror('錯誤', '請先選擇土地與建物謄本 PDF')
            return

        self.btn_run.config(state='disabled')
        threading.Thread(target=self._worker,
                         args=(land_paths, bldg_paths), daemon=True).start()

    def _worker(self, land_paths, bldg_paths):
        try:
            self._log('── 解析謄本 ...')
            data, land = self._combine_parcels(land_paths, bldg_paths)
            data['building_type'] = self.bldg_type_var.get()   # 手動選的建物型態，蓋掉自動判斷
            self._log(f"  地號：{land.get('district','')}{land.get('section','')} {land.get('land_no','')}")
            self._log(f"  門牌：{data.get('address')}")
            if self.data_104:
                self._apply_104(data, self.data_104)
                self._log(f'  ✓ 已套用 104：{self.data_104.get("community_name")}')
            # 查使用分區（無 104，開自己的無頭瀏覽器）
            self._log('── 查詢使用分區 …')
            self._apply_zoning(data, land, driver=None)
            self.after(0, lambda d=data: self._wizard_and_produce(d, close_bot=False))
        except Exception as e:
            import traceback
            self._log(f'❌ 錯誤：{e}')
            self._log(traceback.format_exc())
            self.after(0, lambda: self._reset_after_run(close_bot=False))

    # ── 土地案流程（只需土地謄本，無建物、無 104）──────────
    def _build_land_data(self, land_paths):
        """組土地表 data。座落用地段地號；多筆土地地坪加總。回傳 (data, land0)。"""
        land0 = parse_land(land_paths[0])

        # 地坪 = 面積 × 持分，多筆加總；持分也加總，加起來≈1才算「全部」
        # （同一塊地餵了全部共有人的謄本，地坪加對了但所有權沒有跟著從
        # land0單一筆的「持分」改標「全部」——跟建物案 _combine_parcels 同一個舊 bug）
        total, share_total, got = 0.0, 0.0, False
        for i, lp in enumerate(land_paths):
            li = land0 if i == 0 else parse_land(lp)
            if li.get('land_area') and li.get('land_share'):
                v = _to_ping(li['land_area'] * _share_to_float(li['land_share']))
                self._log(f'   土地{i + 1} 地坪：{v}（持分 {li.get("land_share")}）')
                total += v; got = True
                share_total += _share_to_float(li['land_share'])
        area_land = round(total, 2) if got else None
        if got and len(land_paths) > 1:
            self._log(f'  ✓ 地坪加總 {len(land_paths)} 筆 = {area_land} 坪')
        ownership = '全部' if is_full_ownership(land0.get('land_share')) else '持分'
        if len(land_paths) > 1 and abs(share_total - 1.0) < 0.01:
            ownership = '全部'
            self._log(f'  ✓ 持分加總 ≈ {share_total:.3f}（等於全部），所有權改標「全部」而非「持分」')

        data = {
            'address': f"{land0.get('city','')}{land0.get('district','')}"
                       f"{land0.get('section','')}{land0.get('land_no','')}",
            'area_land': area_land,
            'ownership': ownership,
            'mortgage': bool(land0.get('mortgage_total')),
            'mortgage_amount': land0.get('mortgage_total'),
            'mortgage_bank': land0.get('mortgage_bank'),
            'usage_type': land0.get('usage_type') or '建地',
            'usage_zone_raw': land0.get('usage_zone_raw'),
            # 給使用分區查詢
            'district': land0.get('district', ''),
            'section': land0.get('section', ''),
            'land_no': land0.get('land_no', ''),
            'agent_name': '薛力瑜',
            'selling_points': [],
        }
        return data, land0

    def _worker_land(self, land_paths):
        try:
            self._log('── 解析土地謄本 ...')
            data, land = self._build_land_data(land_paths)
            self._log(f"  地段：{data.get('address')}｜地坪：{data.get('area_land')}｜所有權：{data.get('ownership')}")
            self._log('── 查詢使用分區 …')
            self._apply_zoning(data, land, driver=None)
            if data.get('usage_zone') or data.get('special_zone'):
                self._log(f"  分區：{data.get('special_zone') or data.get('usage_zone')}")
            self.after(0, lambda d=data: self._wizard_and_produce_land(d))
        except Exception as e:
            import traceback
            self._log(f'❌ 錯誤：{e}')
            self._log(traceback.format_exc())
            self.after(0, self._reset_after_run)

    def _wizard_and_produce_land(self, data: dict):
        try:
            self._log('── 跳出土地確認視窗，逐項確認 ...')
            steps = _build_land_steps(is_rental=self._is_rental())
            wiz = ConfirmWizard(self, data, log=self._log, steps=steps)
            result = wiz.run()
            if result is None:
                return
            self._produce_land(result)
        except Exception as e:
            import traceback
            self._log(f'❌ 錯誤：{e}')
            self._log(traceback.format_exc())
        finally:
            self._reset_after_run()

    def _produce_land(self, data: dict):
        is_rental = self._is_rental()
        kind = '土地租賃表' if is_rental else '土地表'
        self._log(f'── 填寫{kind} ...')
        ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        addr_short = (data.get('address') or '土地').replace('高雄市', '').replace('屏東縣', '')[:20]
        out_path = os.path.join(self.out_var.get(), f"{kind}_{addr_short}_{ts}.xlsx")
        fill_land(data, out_path, is_rental=is_rental)
        self._log(f'✅ 完成！輸出：{out_path}')
        self._log('──────────────────────────────')
        if sys.platform == 'win32':
            os.startfile(os.path.dirname(out_path))


if __name__ == '__main__':
    app = App()
    app.mainloop()
